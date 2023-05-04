# =======================================================
# Author : Rizki Nur Rachmadi Yudadiningrat
# SID    : 1900018014
# Subject: Teknik Informatika - Skripsi
# Project: OMRay - scanner for non-LJK exam paper
# Created: 10/11/2022
# Contact: +62 895 0811 7055 (Telegram)
# =======================================================

# import library

import assets.libs.scann as scan
from setting import bukaSetting
from tkinter import messagebox, font, ttk
from tkinter import *
import customtkinter as ctk
import openpyxl as xl
from PIL import Image
import sqlite3
import os

# define root ==========================================

mainMenu = ctk.CTk()
mainMenu.title('OMRay')
mainMenu.geometry('989x732+350+70')
mainMenu.resizable(False, False)
mainMenu.iconbitmap('assets/images/OMRay.ico')

# variable =============================================

xlPath = "assets/temps/omray.xlsx"
ans = [[2, 3, 1, 2, 3, 2, 2, 3, 1, 3],
       [1, 2, 2, 2, 3, 2, 1, 1, 1, 4]]

# header ===============================================

img = ctk.CTkImage(dark_image=Image.open(
    "assets/images/wp1.png"), size=(160, 160))
ctk.CTkLabel(mainMenu, image=img, text=None,bg_color="transparent").place(x=30, y=20)

heading = ctk.CTkLabel(mainMenu, text='OMRay', text_color='#fff', font=('Fugaz One', 48, 'bold'))
heading.place(x=210, y=20)

ctk.CTkFrame(mainMenu, width=740, height=2, fg_color='#fff').place(x=210, y=95)

subheading = ctk.CTkLabel(mainMenu, text='Teacher tool for Exam', text_color='#fff', 
                          bg_color='transparent', font=('Fresca', 24))
subheading.place(x=210, y=105)

subheading2 = ctk.CTkLabel(mainMenu, text='Made by Rizki Nur Rachmadi Y', text_color='#fff', 
                           bg_color='transparent',font=('Fresca', 16))
subheading2.place(x=210, y=135)

# end of header ===========================================

# function ================================================

def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        mainMenu.destroy()

# toplevel
def showSignUp():
    signUpWd = ctk.CTkToplevel(mainMenu)
    signUpWd.title("OMRay | Sign Up")
    signUpWd.geometry('925x500+400+200')
    signUpWd.resizable(False, False)
    signUpWd.iconbitmap('assets/images/OMRay.ico')

    imgbg = ctk.CTkImage(light_image=Image.open("assets/images/bg_wall.png"), size=(1000, 650))
    l1=ctk.CTkLabel(master=signUpWd, image=imgbg, text=None)
    l1.place(x=0, y=0)

    # function ====================================================

    def signup():
        username = user.get()
        password = code.get()
        confirm_password = confirm_code.get()

        if len(username) == 0 or len(password) == 0:
            messagebox.showerror('Invalid', "All fields required!")
        elif password == confirm_password:
            conn.execute("INSERT INTO logins(username, password, status) VALUES (?,?, 'off')",
                         (username, password))
            cursor = conn.execute("SELECT login_id FROM logins WHERE username=? and password=?",(username, password))
            id_long = cursor.fetchone()
            conn.execute("INSERT INTO settings(id_login, cameraNo, showAnswer, autoSave) VALUES (?,0,1,1)",(id_long[0],))
            conn.commit()
            messagebox.showinfo('Sign Up', 'Account created successfully')
            signUpWd.destroy()
            showSignIn()
        else:
            messagebox.showerror('Invalid', "Password doesn't match")

    def signIn_cmd():
        signUpWd.destroy()
        showSignIn()

    # set tampilan ================================================

    frame = ctk.CTkFrame(signUpWd, corner_radius=25, width=350, height=390, bg_color='transparent')
    frame.place(relx=0.5, rely=0.5, anchor=CENTER)

    heading = ctk.CTkLabel(frame, text='SIGN UP', text_color='#FFF', bg_color='transparent',
                    font=('Fugaz One', 28, 'bold'))
    heading.place(relx=0.5, y=40, anchor=CENTER)

    # entry ========================================================

    # username --------------------------------------

    user = ctk.CTkEntry(frame, width=260, height=40, text_color='white', 
                        placeholder_text='Username', bg_color='transparent', font=('Fresca', 18))
    user.place(relx=0.5, y=100, anchor=CENTER)

    # password --------------------------------------

    code = ctk.CTkEntry(frame, width=260, height=40, text_color='white', 
                        placeholder_text='Password', bg_color='transparent', font=('Fresca', 18))
    code.place(relx=0.5, y=160, anchor=CENTER)

    # confirm password -------------------------------

    confirm_code = ctk.CTkEntry(frame, height=40, width=260, text_color='white', 
                                placeholder_text='Confirm Password', bg_color='transparent', 
                                font=('Fresca', 18))
    confirm_code.place(relx=0.5, y=220, anchor=CENTER)

    # button ------------------------------------------

    sign_up = ctk.CTkButton(frame, width=260, height=40, text='Sign Up',
                     bg_color='transparent', corner_radius=6, text_color='white', 
                     cursor='hand2', font=('Fugaz One', 24), command=signup)
    sign_up.pack()
    sign_up.place(relx=0.5, y=290, anchor=CENTER)

    label = ctk.CTkLabel(frame, text="Already have account?", text_color='#fff',
                  bg_color='transparent', font=('Fresca', 16))
    label.place(x=74, y=320)

    sign_in = ctk.CTkButton(frame, width=6, text='Sign In', border_width=0,
                     bg_color='transparent', fg_color="transparent", cursor='hand2', 
                     text_color='#3B92EA', font=('Fresca', 16, 'bold'), command=signIn_cmd)
    sign_in.place(x=220, y=320)

    # Main app---------------------

    signUpWd.mainloop()

# toplevel
def showSignIn():
    signInWd = ctk.CTkToplevel(mainMenu)
    signInWd.title('OMRay | Sign In')
    signInWd.geometry('925x500+400+200')
    signInWd.resizable(False, False)
    signInWd.iconbitmap('assets/images/OMRay.ico')

    # set tampilan ================================================

    imgSi = ctk.CTkImage(light_image=Image.open("assets/images/wp.png"), size=(298, 298))
    ctk.CTkLabel(signInWd, image=imgSi, bg_color='transparent', text=None).place(x=60, rely=0.2)

    frameSignIn = ctk.CTkFrame(signInWd, corner_radius=25, width=350, height=350, bg_color="transparent")
    frameSignIn.place(relx=0.5, y=70)

    heading = ctk.CTkLabel(frameSignIn, text='SIGN IN', text_color='#FFF', bg_color='transparent',
                    font=('Fugaz One', 28, 'bold'))
    heading.place(relx=0.5, y=40, anchor=CENTER)

    # function ====================================================

    def signInCmd():
        username = user.get()
        password = code.get()

        cursor = conn.execute(
            'SELECT * FROM logins WHERE username="%s" and password="%s"' % (username, password))
        if cursor.fetchone():
            conn.execute(
                "UPDATE logins SET status = 'on' WHERE username=? and password=?", (username, password))
            conn.commit()
            mainMenu.deiconify()
            signInWd.destroy()
        else:
            messagebox.showerror("Invalid", "invalid username or password")

    def signUp_cmd():
        signInWd.destroy()
        showSignUp()

    # entry ========================================================

    # username --------------------------------------

    user = ctk.CTkEntry(frameSignIn, height=40, width=260, text_color='white', placeholder_text='Username',
                 bg_color='transparent', font=('Fresca', 18))
    user.place(relx=0.5, y=100, anchor=CENTER)

    # password --------------------------------------

    code = ctk.CTkEntry(frameSignIn, height=40, width=260, text_color='white', placeholder_text='Password',
                        bg_color='transparent', font=('Fresca', 18))
    code.place(relx=0.5, y=160, anchor=CENTER)

    # button ------------------------------------------

    sign_in = ctk.CTkButton(frameSignIn, width=260, height=40, text='Sign In',
                            bg_color='transparent', corner_radius=6, text_color='white', 
                            cursor='hand2',font=('Fugaz One', 24), command=signInCmd)
    sign_in.pack()
    sign_in.place(relx=0.5, y=230, anchor=CENTER)

    label = ctk.CTkLabel(frameSignIn, text="Don't have an account?", text_color='#fff',
                  bg_color='transparent', font=('Fresca', 16))
    label.place(x=72, y=260)

    sign_up = ctk.CTkButton(frameSignIn, width=6, text='Sign Up', border_width=0, 
                            bg_color='transparent', fg_color="transparent", cursor='hand2', 
                            text_color='#3B92EA', font=('Fresca', 16, 'bold'), command=signUp_cmd)
    sign_up.place(x=221, y=260)

    signInWd.protocol("WM_DELETE_WINDOW", on_closing)
    signInWd.mainloop()

def autoRun():
    # creating database
    MYDIR = os.path.dirname(__file__)
    SQLPATH = os.path.join(MYDIR, "assets", "temps", "omr.db")
    global conn
    conn = sqlite3.connect(SQLPATH)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS logins (
        login_id integer primary key autoincrement,
        username text,
        password text,
        status text
    )""")

    conn.execute("""
    CREATE TABLE IF NOT EXISTS subjects (
        sub_id integer primary key autoincrement,
        sub_name text,
        sub_totalQuestion integer,
        sub_choices integer,
        sub_answer text
    )""")
    
    conn.execute("""
    CREATE TABLE IF NOT EXISTS settings (
        set_id integer primary key autoincrement,
        id_login integer, 
        cameraNo integer, 
        def_subject text, 
        showAnswer integer, 
        autoSave integer
    )""")

    # conn.execute("INSERT INTO subjects(sub_name, sub_totalQuestion, sub_choices, sub_answer) VALUES ('Matematika', 20, 5, '2, 3, 1, 2, 3, 2, 2, 3, 1, 3, 1, 2, 2, 2, 3, 2, 1, 1, 1, 4')")

    conn.commit()

    cursor = conn.execute(
        'SELECT * FROM logins WHERE status= "on"')
    if cursor.fetchone():
        mainMenu.deiconify()
    else:
        # hide dashboard ---------
        mainMenu.withdraw()
        showSignIn()

def logout():
    mainMenu.withdraw()
    conn.execute("UPDATE logins SET status = 'off' WHERE status = 'on'")
    conn.commit()
    showSignIn()

# toplevel
def setting():
    print ("test")

def test():
    scan.newScanning(True, 0, 10, 5, ans, 0)

# end of function =========================================

# buttons =================================================

# Logout button
img_logout = ctk.CTkImage(light_image=Image.open("assets/images/logout.png"), size=(20, 20))

logout_button = ctk.CTkButton(mainMenu, text="Logout", image=img_logout, compound=RIGHT, 
                              width=30, height=30, corner_radius=100, cursor='hand2', 
                              font=('Fredoka One', 14, 'bold'), command=logout)
logout_button.place(x=835, y=20)

# add subject
img_addsbj = ctk.CTkImage(light_image=Image.open(
    "assets/images/add_sbj.png"), size=(90, 90))

addsbj_button = ctk.CTkButton(mainMenu, text="Subject", image=img_addsbj, compound=TOP, 
                              fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                              corner_radius=25, width=200, height=180, cursor='hand2', 
                              font=('Fredoka One', 24, 'bold'))
addsbj_button.place(x=30, y=210)

# add record
img_addrec = ctk.CTkImage(light_image=Image.open(
    "assets/images/add_rec.png"), size=(90, 90))

addrec_button = ctk.CTkButton(mainMenu, text="Record", image=img_addrec, compound=TOP, 
                              fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                              corner_radius=25, width=200, height=180, cursor='hand2', 
                              font=('Fredoka One', 24, 'bold'))
addrec_button.place(x=250, y=210)

# setting
img_sett = ctk.CTkImage(light_image=Image.open(
    "assets/images/sett.png"), size=(90, 90))

sett_button = ctk.CTkButton(mainMenu, text="Setting", image=img_sett, compound=TOP, 
                            fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                            corner_radius=25, width=200, height=180, cursor='hand2', 
                            font=('Fredoka One', 24, 'bold'), command=bukaSetting)
sett_button.place(x=30, y=410)

# see record
img_seerec = ctk.CTkImage(light_image=Image.open(
    "assets/images/see_exc.png"), size=(85, 85))

seerec_button = ctk.CTkButton(mainMenu, text="See Record", image=img_seerec, compound=TOP, 
                              fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                              corner_radius=25, width=200, height=180, cursor='hand2', 
                              font=('Fredoka One', 24, 'bold'))
seerec_button.place(x=250, y=410)

# Create a new scan
img_scan = ctk.CTkImage(light_image=Image.open(
    "assets/images/scann.png"), size=(60, 60))

scan_button = ctk.CTkButton(mainMenu, text="Create a New Scan", image=img_scan, compound=LEFT, 
                            fg_color="#1798F8", hover_color="#085288", text_color="#fff", 
                            corner_radius=25, width=400, height=100, cursor='hand2', 
                            font=('Fredoka One', 32, 'bold'), command=test)
scan_button.place(x=30, y=610)

# end of button ===========================================

# scrollable frame ========================================

scrollable_frame = ctk.CTkScrollableFrame(mainMenu, width=430, height=560)
scrollable_frame.place(x=500, y=135)

title_label = ctk.CTkLabel(scrollable_frame, text="Scanning History",
                           font=('Fugaz One', 42, 'bold'))
title_label.pack(fill="x")

ctk.CTkFrame(scrollable_frame, width=740, height=2,
             fg_color='#fff').pack(padx=5, pady=10)

# function load data excel

def loadExcel():
    if not os.path.exists(xlPath):
        workbook = xl.Workbook()

        sheet0 = workbook.active
        sheet0.title = "Scanning history"
        heading0 = ["Date/time", "Subject", "Classroom"]
        sheet0.append(heading0)

        workbook1 = workbook.create_sheet("Sheet_A")
        sheet1 = workbook1
        sheet1.title = "Scanning result"
        heading1 = ["Date/time", "Subject", "Classroom",
                    "Student ID", "Grade", "Wrong Answer"]
        sheet1.append(heading1)

        workbook.save(xlPath)
    
    workbook = xl.load_workbook(xlPath)
    sheet = workbook.active

    list_data = list(sheet.values)
    cols = list_data[0]

    tree = ttk.Treeview(scrollable_frame, columns=cols, show="headings", height=27)
    for col_name in cols:
        tree.heading(col_name, text=col_name)
        tree.column(col_name, width=font.Font().measure(col_name))

    tree.pack(expand=True, fill="both")

    for value_data in list_data[1:]:
        tree.insert('', 'end', values=value_data)

    return tree

tree = loadExcel()

def update_treeview():
    workbook = xl.load_workbook(xlPath)
    
    workbook._active_sheet_index = 0
    sheet = workbook.active

    list_data = list(sheet.values)
    cols = list_data[0]

    tree.delete(*tree.get_children())

    tree.config(columns=cols)
    for col_name in cols:
        tree.heading(col_name, text=col_name)
        tree.column(col_name, width=font.Font().measure(col_name))

    for value_data in list_data[1:]:
        tree.insert("", 'end', values=value_data)

    mainMenu.after(1000, update_treeview)

# end of scrollable frame ==================================

mainMenu.protocol("WM_DELETE_WINDOW", on_closing)
autoRun()
update_treeview()

mainMenu.mainloop()
