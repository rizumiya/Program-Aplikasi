# =======================================================
# Author : Rizki Nur Rachmadi Yudadiningrat
# SID    : 1900018014
# Subject: Teknik Informatika - Skripsi
# Project: OMRay - scanner for non-LJK exam paper
# Created: 10/11/2022
# Contact: +62 895 0811 7055 (Telegram)
# =======================================================


from tkinter import messagebox, font, ttk
from tkinter import *
import customtkinter as ctk
import openpyxl as xl
from PIL import Image
import sqlite3
import os

from assets.libs.Database import Database
from assets.libs.Module import Module

class MainMenu(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title('OMRay')
        self.geometry('989x732+60+65')
        self.resizable(False, False)
        self.iconbitmap('assets/images/OMRay.ico')

        self.xlPath = "assets/temps/omray.xlsx"

        # creating database
        
        self.SQLPATH = os.path.join(os.path.dirname(__file__), 'assets', 'temps', 'omr.db')
        global conn
        self.conn = sqlite3.connect(self.SQLPATH)
        self.cur = self.conn.cursor()

        # variable

        self.database = Database()
        self.my_module = Module()
        self.camNo = str
        self.subj = str
        self.shwAns = str
        self.autoSv = str

        # header ===============================================

        self.img = ctk.CTkImage(dark_image=Image.open("assets/images/wp1.png"), size=(160, 160))
        ctk.CTkLabel(self, image=self.img, text=None,bg_color="transparent").place(x=30, y=20)

        self.heading = ctk.CTkLabel(self, text='OMRay', text_color='#fff', font=('Fugaz One', 48, 'bold'))
        self.heading.place(x=210, y=20)

        ctk.CTkFrame(self, width=740, height=2, fg_color='#fff').place(x=210, y=95)

        self.subheading = ctk.CTkLabel(self, text='Teacher tool for Exam', text_color='#fff', 
                                       bg_color='transparent', font=('Fresca', 24))
        self.subheading.place(x=210, y=105)

        self.subheading2 = ctk.CTkLabel(self, text='Made by Rizki Nur Rachmadi Y', text_color='#fff',
                                        bg_color='transparent',font=('Fresca', 16))
        self.subheading2.place(x=210, y=135)

        # buttons =================================================

        # Logout
        self.img_logout = ctk.CTkImage(light_image=Image.open("assets/images/logout.png"), size=(20, 20))

        self.logout_button = ctk.CTkButton(self, text="Logout", image=self.img_logout, compound=RIGHT, 
                                    width=30, height=30, corner_radius=100, cursor='hand2', 
                                    font=('Fredoka One', 14, 'bold'), command=self.logout)
        self.logout_button.place(x=835, y=20)

        # add subject
        self.img_addsbj = ctk.CTkImage(light_image=Image.open(
            "assets/images/add_sbj.png"), size=(90, 90))

        self.addsbj_button = ctk.CTkButton(self, text="Subject", image=self.img_addsbj, compound=TOP, 
                                    fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                                    corner_radius=25, width=200, height=180, cursor='hand2', 
                                    font=('Fredoka One', 24, 'bold'))
        self.addsbj_button.place(x=30, y=210)

        # add record
        self.img_addrec = ctk.CTkImage(light_image=Image.open(
            "assets/images/add_rec.png"), size=(90, 90))

        self.addrec_button = ctk.CTkButton(self, text="Record", image=self.img_addrec, compound=TOP, 
                                    fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                                    corner_radius=25, width=200, height=180, cursor='hand2', 
                                    font=('Fredoka One', 24, 'bold'))
        self.addrec_button.place(x=250, y=210)

        # setting
        self.img_sett = ctk.CTkImage(light_image=Image.open(
            "assets/images/sett.png"), size=(90, 90))

        self.sett_button = ctk.CTkButton(self, text="Setting", image=self.img_sett, compound=TOP, 
                                    fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                                    corner_radius=25, width=200, height=180, cursor='hand2', 
                                    font=('Fredoka One', 24, 'bold'), command=self.bukaSetting)
        self.sett_button.place(x=30, y=410)

        # see record
        self.img_seerec = ctk.CTkImage(light_image=Image.open(
            "assets/images/see_exc.png"), size=(85, 85))

        self.seerec_button = ctk.CTkButton(self, text="See Record", image=self.img_seerec, compound=TOP, 
                                    fg_color="#fff", hover_color="#B2B2B2", text_color="#333030", 
                                    corner_radius=25, width=200, height=180, cursor='hand2', 
                                    font=('Fredoka One', 24, 'bold'))
        self.seerec_button.place(x=250, y=410)

        # Create a new scan
        self.img_scan = ctk.CTkImage(light_image=Image.open(
            "assets/images/scann.png"), size=(60, 60))

        self.scan_button = ctk.CTkButton(self, text="Create a New Scan", image=self.img_scan, compound=LEFT, 
                                    fg_color="#1798F8", hover_color="#085288", text_color="#fff", 
                                    corner_radius=25, width=400, height=100, cursor='hand2', 
                                    font=('Fredoka One', 32, 'bold'), command=self.quickScan)
        self.scan_button.place(x=30, y=610)

        # scrollable frame ========================================

        self.scrollable_frame = ctk.CTkScrollableFrame(self, width=430, height=560)
        self.scrollable_frame.place(x=500, y=135)

        self.title_label = ctk.CTkLabel(self.scrollable_frame, text="Scanning History",
                                font=('Fugaz One', 42, 'bold'))
        self.title_label.pack(fill="x")

        ctk.CTkFrame(self.scrollable_frame, width=740, height=2,
                    fg_color='#fff').pack(padx=5, pady=10)

        self.tree = self.loadExcel()
        self.ambilSetting()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    # function 

    def loadExcel(self):
        if not os.path.exists(self.xlPath):
            self.workbook = xl.Workbook()

            self.sheet0 = self.workbook.active
            self.sheet0.title = "Scanning history"
            self.heading0 = ["Date/time", "Subject", "Classroom"]
            self.sheet0.append(self.heading0)

            self.workbook1 = self.workbook.create_sheet("Sheet_A")
            self.sheet1 = self.workbook1

            self.sheet1.title = "Scanning result"
            self.heading1 = ["Date/time", "Subject", "Classroom",
                        "Student ID", "Grade", "Wrong Answer"]
            self.sheet1.append(self.heading1)
            self.workbook.save(self.xlPath)
        
        self.workbook = xl.load_workbook(self.xlPath)
        self.sheet = self.workbook.active
        self.list_data = list(self.sheet.values)
        self.cols = self.list_data[0]

        self.tree = ttk.Treeview(self.scrollable_frame, columns=self.cols, show="headings", height=27)
        for self.col_name in self.cols:
            self.tree.heading(self.col_name, text=self.col_name)
            self.tree.column(self.col_name, width=font.Font().measure(self.col_name))
        self.tree.pack(expand=True, fill="both")

        for self.value_data in self.list_data[1:]:
            self.tree.insert('', 'end', values=self.value_data)
        return self.tree


    def update_treeview(self):
        self.workbook = xl.load_workbook(self.xlPath)
        
        self.workbook._active_sheet_index = 0
        self.sheet = self.workbook.active
        self.list_data = list(self.sheet.values)
        self.cols = self.list_data[0]
        self.tree.delete(*self.tree.get_children())
        self.tree.config(columns=self.cols)

        for col_name in self.cols:
            self.tree.heading(col_name, text=col_name)
            self.tree.column(col_name, width=font.Font().measure(col_name))

        for value_data in self.list_data[1:]:
            self.tree.insert("", 'end', values=value_data)
        self.after(1000, self.update_treeview)

        # end of scrollable frame ==================================

        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.update_treeview()


    def autoRun(self):
        self.database.createTables()

        self.cursor = self.conn.execute('SELECT * FROM logins WHERE status= "on"')
        if self.cursor.fetchone():
            self.deiconify()
        else:
            self.destroy()
            from pages.LoginForm import LoginForm
            login_form = LoginForm()
            login_form.mainloop()


    def logout(self):
        self.conn.execute("UPDATE logins SET status = 'off' WHERE status = 'on'")
        self.conn.commit()
        self.destroy()
        from pages.LoginForm import LoginForm
        login_form = LoginForm()
        login_form.mainloop()


    def bukaSetting(self):
        from pages.SettingPage import SettingPage
        setting_page = SettingPage(self, "OMRay | Setting")
        setting_page.mainloop()


    def bukaSubject(self):
        self.destroy()
        from pages.SubjectPage import SubjectPage
        subject_page = SubjectPage()
        subject_page.mainloop()


    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.destroy()


    def ambilSetting(self):
        self.id_login = self.database.selectActive()
        self.dataSetting = self.database.selectAttributes('*', 'settings', 'id_login=' + str(self.id_login))
        self.camNo = self.dataSetting[0][2]
        self.subj = self.dataSetting[0][3]
        self.shwAns = self.dataSetting[0][4]
        self.autoSv = self.dataSetting[0][5]


    def quickScan(self):
        ansid = 0
        ans = [[2, 3, 1, 2, 3, 2, 2, 3, 1, 3],
            [1, 2, 2, 2, 3, 2, 1, 1, 1, 4]]
        
        from assets.libs.scann import newScanning
        newScanning(True, int(self.camNo), 10, 5, ans, ansid)



if __name__ == '__main__':
    app = MainMenu()
    app.autoRun()
    app.mainloop()